import io
import os
import pandas as pd
from openpyxl import load_workbook
from data_process.generate_logger import *


# 特征（变量）类型初步判定
def check_feature_type(df):

    features_type = {"Ordered": [], "Classification": [], "Continuous": [], "Other": []}

    for column in df.columns:
        data_type = df[column].dtype
        # 根据数据类型选择绘图方案
        if data_type == 'int64':
            features_type["Ordered"].append(column)
        elif data_type == 'float64':
            if df[column].nunique()/df.shape[0] < 0.01:
                features_type["Ordered"].append(column)
            else:
                features_type["Continuous"].append(column)
        elif data_type == 'object':
            features_type["Classification"].append(column)
        else:
            features_type["Other"].append(column)

    return features_type["Ordered"], features_type["Classification"], features_type["Continuous"], features_type["Other"]


def data_check_summary(df, sheet_name):
    """
    获取DataFrame对象中包含缺失值的行，并将结果写入Excel工作表

    参数：
    df: DataFrame对象
    output_file: 输出的Excel文件路径
    sheet_name: 工作表名称

    返回值：
    DataFrame对象中包含缺失值的行
    """
    rows_with_missing_values = df[df.isnull().any(axis=1)]
    data_analyze = analyze_dataframe(df)  # Implement this function separately
    data_summary_statistics = df.describe()
    data_line5 = df.head(5)

    output_file = r"datainfo_check.xlsx"

    if not os.path.isfile(output_file):
        # Create a new workbook
        writer = pd.ExcelWriter(output_file, engine='openpyxl')
        data_line5.to_excel(writer, sheet_name=sheet_name, index=False)
        data_analyze.to_excel(writer, sheet_name=sheet_name, startrow=len(data_line5) + 2, index=True)
        data_summary_statistics.to_excel(writer, sheet_name=sheet_name,
                                         startrow=len(data_line5) + len(data_analyze) + 4, index=True)
        rows_with_missing_values.to_excel(writer, sheet_name=sheet_name,
                                          startrow=len(data_line5) + len(data_analyze) + len(
                                              data_summary_statistics) + 6, index=False)
        writer.save()
    else:
        # 检查子表格是否存在
        book = load_workbook(output_file)
        if sheet_name in book.sheetnames:
            # 如果子表格存在，则删除它
            book.remove(book[sheet_name])
            book.save(output_file)
        # Append to an existing workbook
        with pd.ExcelWriter(output_file, mode='a', engine='openpyxl') as writer:
            data_line5.to_excel(writer, sheet_name=sheet_name, index=False)
            data_analyze.to_excel(writer, sheet_name=sheet_name, startrow=len(data_line5) + 2, index=True)
            data_summary_statistics.to_excel(writer, sheet_name=sheet_name,
                                             startrow=len(data_line5) + len(data_analyze) + 4, index=True)
            rows_with_missing_values.to_excel(writer, sheet_name=sheet_name,
                                              startrow=len(data_line5) + len(data_analyze) + len(
                                                  data_summary_statistics) + 6, index=False)
    return rows_with_missing_values


# 计算每一列缺失值的占比情况
def missing_values(df):
    """
    计算每列的缺失值及占比
    """
    missing_values_count = df.isnull().sum().sort_values(ascending=False)
    missing_values_percent = (df.isnull().sum() / df.isnull().count()).sort_values(ascending=False)
    missing_values_df = pd.concat([missing_values_count, missing_values_percent], axis=1)
    missing_values_df.columns = (['Missing_Values_Count', 'Missing_Values_Percent'])
    return missing_values_df


def analyze_dataframe(df):
    # 调用missing_values()方法获取缺失值信息
    missing_values_df = missing_values(df)

    # 合并结果
    result = pd.concat([df[df.columns].nunique(), df[df.columns].dtypes, missing_values_df], axis=1)
    result.columns = ['Unique Values', 'Data Types', 'Missing Values Count', 'Missing Values Percent']

    return result


# 检测CSV表格中哪些列数据元素是唯一的
def get_unique_columns(df):
    unique_columns = []

    for column_name in df.columns:
        column_data = df[column_name]
        is_unique = column_data.is_unique

        if is_unique:
            unique_columns.append(column_name)

    return unique_columns


# 检查数据集训练集和测试集分布是否一致
def check_data_info(data_path, sheet_name=""):
    logfile_path = "./dataset_info.log"
    logger = configure_logger(logfile_path)
    # 设置显示所有列
    pd.set_option('display.max_columns', None)

    if sheet_name == "train":
        logger.info("*****************《{}》数据核查*****************".format(sheet_name))
        data = pd.read_csv(data_path)
        logger.info("训练集所有列索引名称：\n {}".format(data.columns))
        logger.info("训练集特征（变量）数据类型初步判定：\n ClassifyOrdered:{} \n Classification:{} \n Continuous:{} \n Other:{}"
                    .format(*check_feature_type(data)))
        logger.info("训练集列元素是唯一的列索引名称：{}".format(get_unique_columns(data)))
        logger.info("训练集总行数和总列数：{}".format(data.shape))
        # 将data.info()的结果写入日志文件
        with io.StringIO() as buf:
            data.info(buf=buf)
            buf.seek(0)
            logger.info(buf.read())
        # 将总结的数据信息写入Excel文件
        data_check_summary(data, sheet_name)
    elif sheet_name == "test":
        logger.info("*****************《{}》数据核查*****************".format(sheet_name))
        data = pd.read_csv(data_path)
        logger.info("测试集所有列索引名称：\n {}".format(data.columns))
        logger.info("测试集特征（变量）数据类型初步判定：\n Ordered:{} \n Classification:{} \n Continuous:{} \n Other:{}"
                    .format(*check_feature_type(data)))
        logger.info("测试集列元素是唯一的列索引名称：{}".format(get_unique_columns(data)))
        logger.info("测试集总行数和总列数：{}".format(data.shape))
        # 将data.info()的结果写入日志文件
        with io.StringIO() as buf:
            data.info(buf=buf)
            buf.seek(0)
            logger.info(buf.read())
        # 将总结的数据信息写入Excel文件
        data_check_summary(data, sheet_name)
    else:
        logger.info("*****************《{}》数据集核查*****************".format(sheet_name))
        data = pd.read_csv(data_path)
        logger.info("数据集所有列索引名称：\n {}".format(data.columns))
        logger.info("数据集特征（变量）数据类型初步判定：\n Ordered:{} \n Classification:{} \n Continuous:{} \n Other:{}"
                    .format(*check_feature_type(data)))
        logger.info("数据集列元素是唯一的列索引名称：{}".format(get_unique_columns(data)))
        logger.info("数据集总行数和总列数：{}".format(data.shape))
        # 将data.info()的结果写入日志文件
        with io.StringIO() as buf:
            data.info(buf=buf)
            buf.seek(0)
            logger.info(buf.read())
        # 将总结的数据信息写入Excel文件
        data_check_summary(data, sheet_name)


check_data_info(r"/home/leon/gitlocal/ml_code/ori_dataset/merchants.csv", 'merchants')
